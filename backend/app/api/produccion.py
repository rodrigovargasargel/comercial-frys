from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import io #para excel con formato bonito al igual que las openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as date_type
from sqlalchemy.orm import Session
from typing import List
from app.db.database import get_db
from app.models.producto import TipoProducto, Producto
from app.models.produccion import Color
from app.schemas.produccion import (
    OrdenProduccionCreate, OrdenProduccionUpdate, OrdenProduccionOut,
    ProduccionExtrusoraCreate, ProduccionExtrusoraOut,
    DetalleCreate, DetalleOut,
    TipoProductoSimple, ColorSimple, ProductoSimpleOut
)
from app.services import produccion_service

router = APIRouter(prefix="/produccion", tags=["Produccion"])

@router.get("/tipos-producto", response_model=List[TipoProductoSimple])
def listar_tipos_producto(db: Session = Depends(get_db)):
    return db.query(TipoProducto).all()

@router.get("/colores", response_model=List[ColorSimple])
def listar_colores(db: Session = Depends(get_db)):
    return db.query(Color).all()




@router.get("/productos-extrusora", response_model=List[ProductoSimpleOut])
def listar_productos_extrusora(db: Session = Depends(get_db)):
    from app.models.maquina import TipoMaquina
    tipo_extrusora = db.query(TipoMaquina).filter(TipoMaquina.nombre == "Extrusora").first()
    if not tipo_extrusora:
        return []
    return db.query(Producto).filter(Producto.tipo_maquina_id == tipo_extrusora.id).all()


@router.get("/ops", response_model=List[OrdenProduccionOut])
def listar_ops(db: Session = Depends(get_db)):
    return produccion_service.get_all_ops(db)

@router.get("/ops/{op_id}", response_model=OrdenProduccionOut)
def obtener_op(op_id: int, db: Session = Depends(get_db)):
    op = produccion_service.get_op_by_id(db, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="OP no encontrada")
    return op

@router.post("/ops", response_model=OrdenProduccionOut, status_code=201)
def crear_op(data: OrdenProduccionCreate, db: Session = Depends(get_db)):
    return produccion_service.create_op(db, data)

#Editar turno extrusora
@router.put("/producciones/{produccion_id}", response_model=ProduccionExtrusoraOut)
def actualizar_produccion(produccion_id: int, data: dict, db: Session = Depends(get_db)):
    from app.models.produccion import ProduccionExtrusora
    prod = db.query(ProduccionExtrusora).filter(ProduccionExtrusora.id == produccion_id).first()
    if not prod:
        raise HTTPException(status_code=404, detail="No encontrado")
    for field in ['maquina_id', 'turno', 'lote', 'usuario_id']:
        if field in data:
            setattr(prod, field, data[field])
    db.commit()
    db.refresh(prod)
    return produccion_service._enrich_produccion(db, prod)


@router.put("/ops/{op_id}", response_model=OrdenProduccionOut)
def actualizar_op(op_id: int, data: OrdenProduccionUpdate, db: Session = Depends(get_db)):
    op = produccion_service.update_op(db, op_id, data)
    if not op:
        raise HTTPException(status_code=404, detail="OP no encontrada")
    return op

@router.delete("/ops/{op_id}", status_code=204)
def eliminar_op(op_id: int, db: Session = Depends(get_db)):
    if not produccion_service.delete_op(db, op_id):
        raise HTTPException(status_code=404, detail="OP no encontrada")

@router.get("/ops/{op_id}/producciones", response_model=List[ProduccionExtrusoraOut])
def listar_producciones(op_id: int, db: Session = Depends(get_db)):
    return produccion_service.get_producciones_by_op(db, op_id)

@router.post("/producciones", response_model=ProduccionExtrusoraOut, status_code=201)
def crear_produccion(data: ProduccionExtrusoraCreate, db: Session = Depends(get_db)):
    return produccion_service.create_produccion(db, data)

@router.delete("/producciones/{produccion_id}", status_code=204)
def eliminar_produccion(produccion_id: int, db: Session = Depends(get_db)):
    if not produccion_service.delete_produccion(db, produccion_id):
        raise HTTPException(status_code=404, detail="Producción no encontrada")

@router.get("/producciones/{produccion_id}/detalles", response_model=List[DetalleOut])
def listar_detalles(produccion_id: int, db: Session = Depends(get_db)):
    return produccion_service.get_detalles_by_produccion(db, produccion_id)

@router.post("/detalles", response_model=DetalleOut, status_code=201)
def crear_detalle(data: DetalleCreate, db: Session = Depends(get_db)):
    try:
        return produccion_service.create_detalle(db, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/detalles/{detalle_id}", status_code=204)
def eliminar_detalle(detalle_id: int, db: Session = Depends(get_db)):
    if not produccion_service.delete_detalle(db, detalle_id):
        raise HTTPException(status_code=404, detail="Detalle no encontrado")
    
@router.get("/ocs")
def listar_ocs(empresa_id: int = None, db: Session = Depends(get_db)):
    from app.models.produccion import OrdenProduccion
    query = db.query(OrdenProduccion.oc_cliente)\
        .filter(OrdenProduccion.oc_cliente != None)\
        .filter(OrdenProduccion.oc_cliente != '')
    if empresa_id:
        query = query.filter(OrdenProduccion.empresa_id == empresa_id)
    ocs = query.distinct().all()
    return [oc[0] for oc in ocs]   
    
#excel con formato bonito

@router.post("/ops/{op_id}/packing")

def generar_packing(op_id: int, body: dict, db: Session = Depends(get_db)):
    from datetime import date as date_type
    op = produccion_service.get_op_by_id(db, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="OP no encontrada")
    hoy = date_type.today()
    fecha_str = hoy.strftime('%d/%m/%Y')  # ← debe estar antes
    cliente = body.get('cliente', '')
    ref = body.get('ref', '')
    fact = body.get('fact', '')
    fecha_fact_str = body.get('fecha_fact', fecha_str)

    # Formatear fecha_fact si viene como YYYY-MM-DD
    if fecha_fact_str and '-' in fecha_fact_str:
         partes = fecha_fact_str.split('-')
    fecha_fact_str = f"{partes[2]}/{partes[1]}/{partes[0]}"

    prods = produccion_service.get_producciones_by_op(db, op_id)
    from app.models.produccion import DetalleProduccionExtrusora
    
    thin = Side(style='thin')
    medium = Side(style='medium')
    border_full = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1F3864')
    subheader_fill = PatternFill('solid', fgColor='2E75B6')
    row_fill_alt = PatternFill('solid', fgColor='EBF3FB')

    wb = Workbook()
    ws = wb.active
    ws.title = str(fact)

    anchos = [3, 10, 14, 40, 10, 14, 8, 10]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    hoy = date_type.today()
    fecha_str = hoy.strftime('%d/%m/%Y')

    # Nº

    hoy = date_type.today()
    numero_packing = f"{hoy.strftime('%m%d%y')}" 
   
    ws['G2'] = 'Nº DOC :'
    ws['G2'].font = Font(bold=True, size=10, color='1F3864')
    ws['I2'] = numero_packing
    ws['I2'].font = Font(bold=True, size=12, color='1F3864')
    ws['I2'].alignment = Alignment(horizontal='center')

    ws['G3'] = 'NP :'
    ws['G3'].font = Font(bold=True, size=10, color='1F3864')
    ws['I3'] = op_id
    ws['I3'].font = Font(bold=True, size=12, color='1F3864')
    ws['I3'].alignment = Alignment(horizontal='center')

    # Empresa
    empresa_rows = [
        (4, 'COMERCIALIZADORA Y DISTRIBUIDORA FRYS LTDA', True, 12),
        (5, 'RUT : 76386703-K', False, 10),
        (6, 'CAMINO TEPUAL km 3 , PARQUE APIASMONT PARCELA 9', False, 10),
        (7, 'PUERTO MONTT - CHILE', False, 10),
        (8, 'TEL (065) 2713834 - (09) 96797817', False, 10),
        (9, 'GIRO : COMERCIALIZACION DE INSUMOS Y PRODUCTOS INDUSTRIALES', False, 10),
        (10, 'rmolina@comercialfrys.cl', False, 10),
    ]
    for row, val, bold, size in empresa_rows:
        c = ws[f'B{row}']
        c.value = val
        c.font = Font(bold=bold, size=size, color='1F3864')

    # Título
    ws.merge_cells('B12:I13')
    c = ws['B12']
    c.value = 'PACKING LIST'
    c.font = Font(bold=True, size=24, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = header_fill
    ws.row_dimensions[12].height = 20
    ws.row_dimensions[13].height = 20

    # Separador
    for col in range(2, 10):
        ws.cell(row=14, column=col).fill = PatternFill('solid', fgColor='2E75B6')
    ws.row_dimensions[14].height = 4

    # A/TO y fecha
    ws['B16'] = 'A/TO :'
    ws['B16'].font = Font(bold=True, size=10, color='1F3864')
    ws['D16'] = cliente.upper()
    ws['D16'].font = Font(bold=True, size=11)
    ws['G16'] = 'FECHA/DATE :'
    ws['G16'].font = Font(bold=True, size=10, color='1F3864')
    ws['G18'] = 'NOTA PEDIDO :'
    ws['G18'].font = Font(bold=True, size=10, color='1F3864')
    ws['I16'] = fecha_str
    ws['I16'].font = Font(size=10)
    ws['I16'].alignment = Alignment(horizontal='center')

    ws['I18'] = op_id
    ws['I18'].font = Font(size=10)
    ws['I18'].alignment = Alignment(horizontal='center')


    ws['B18'] = 'REF. :'
    ws['B18'].font = Font(bold=True, size=10, color='1F3864')
    ws['D18'] = ref
    ws['D18'].font = Font(size=10)

    # Encabezado tabla
    headers = [('B','FACT'),('C','FECHA'),('D','NOMBRE PRODUCTO'),
           ('F','LOTE'),('G','FECHA PROD.'),('H','ROLLO'),('I','KG')]
    HEADER_ROW = 21
    ws.row_dimensions[HEADER_ROW].height = 18
    for col, h in headers:
        c = ws[f'{col}{HEADER_ROW}']
        c.value = h
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = subheader_fill
        c.border = border_full
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Merge D y E para NOMBRE PRODUCTO
    ws.merge_cells(f'D{HEADER_ROW}:E{HEADER_ROW}')
    e_header = ws[f'E{HEADER_ROW}']
    e_header.fill = subheader_fill
    e_header.border = Border(
        top=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style=None)  # sin borde izquierdo
    )
   

    # Datos
    #nombre_producto = f"{op.producto.nombre if op.producto else ''} {op.color.nombre if op.color else ''} {op.ancho}x{op.espesor} {op.densidad}"
    densidad_str = 'AD' if op.densidad == 'alta' else 'BD'
    nombre_producto = f"{op.producto.nombre if op.producto else ''} {densidad_str} {op.color.nombre if op.color else ''} {op.ancho}x{op.espesor} "
    DATA_START = HEADER_ROW + 1
    fila = DATA_START
    primera = True

    for prod in prods:
        dets = db.query(DetalleProduccionExtrusora)\
            .filter(DetalleProduccionExtrusora.produccion_extrusora_id == prod.id)\
            .order_by(DetalleProduccionExtrusora.numero_rollo).all()
        fecha_prod = prod.fecha.strftime('%d/%m/%Y') if prod.fecha else ''
        for det in dets:
            fill = PatternFill('solid', fgColor='FFFFFF') if fila % 2 == 0 else row_fill_alt
            ws.row_dimensions[fila].height = 15
            valores = [
                ('B', int(fact) if str(fact).isdigit() else fact),
                ('C', fecha_fact_str),
                ('D', nombre_producto),
                ('F', prod.lote),
                ('G', fecha_prod),
                ('H', det.numero_rollo if primera else f'=H{fila-1}+1'),
                ('I', det.kg),
            ]
            for col, val in valores:
                c = ws[f'{col}{fila}']
                c.value = val
                c.font = Font(size=10)
                c.border = border_full
                c.fill = fill
                if col in ('H', 'I'):
                    c.alignment = Alignment(horizontal='center')
            primera = False
            fila += 1
            ws.merge_cells(f'D{fila}:E{fila}')
             # Aplicar formato a celda E también
            e_cell = ws[f'E{fila}']
            e_cell.fill = fill
            e_cell.border = Border(
                top=Side(style='thin'),
                right=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style=None)  # sin borde izquierdo
            )
            

    # Total
    total_row = fila
    ws.row_dimensions[total_row].height = 18
    ws.merge_cells(f'B{total_row}:G{total_row}')
    c = ws[f'B{total_row}']
    c.value = 'TOTAL'
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = header_fill
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = border_full
    ws.merge_cells(f'B{total_row}:H{total_row}')
    # Agregar formato a H del total
    h_total = ws[f'H{total_row}']
    h_total.fill = header_fill
    h_total.border = border_full
    tc = ws[f'I{total_row}']
    tc.value = f'=SUM(I{DATA_START}:I{fila-1})'
    tc.font = Font(bold=True, size=11, color='FFFFFF')
    tc.fill = header_fill
    tc.border = border_full
    tc.alignment = Alignment(horizontal='center', vertical='center')

    # Firma
    firma_row = total_row + 5
    for i, linea in enumerate(['Renzo Molina E.', 'Dpto. Comercial', 'F : 09-96797817', 'rmolina@comercialfrys.cl']):
        c = ws[f'B{firma_row+i}']
        c.value = linea
        c.font = Font(size=10, color='1F3864', italic=(i > 0))

    # Stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Packing-{fact}-{cliente}.xlsx"
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )    


