from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from app.db.database import get_db
from app.models.oc_ventas import OCVenta
from app.schemas.oc_ventas import OCVentaCreate, OCVentaOut
import io

router = APIRouter(prefix="/oc-ventas", tags=["oc_ventas"])

@router.get("/", response_model=List[OCVentaOut])
def listar(db: Session = Depends(get_db)):
    return db.query(OCVenta).order_by(OCVenta.fecha.desc()).all()

@router.post("/", response_model=OCVentaOut, status_code=201)
def crear(data: OCVentaCreate, db: Session = Depends(get_db)):
    oc = OCVenta(**data.dict())
    db.add(oc)
    db.commit()
    db.refresh(oc)
    return oc

@router.put("/{id}", response_model=OCVentaOut)
def actualizar(id: int, data: OCVentaCreate, db: Session = Depends(get_db)):
    oc = db.query(OCVenta).filter(OCVenta.id == id).first()
    if not oc:
        raise HTTPException(status_code=404, detail="No encontrado")
    for k, v in data.dict().items():
        setattr(oc, k, v)
    db.commit()
    db.refresh(oc)
    return oc

@router.delete("/{id}")
def eliminar(id: int, db: Session = Depends(get_db)):
    oc = db.query(OCVenta).filter(OCVenta.id == id).first()
    if not oc:
        raise HTTPException(status_code=404, detail="No encontrado")
    db.delete(oc)
    db.commit()
    return {"ok": True}

@router.get("/excel")
def exportar_excel(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ocs = db.query(OCVenta).order_by(OCVenta.fecha.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "OC Ventas"

    header_fill = PatternFill('solid', fgColor='1F3864')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['ID', 'Fecha', 'Cliente', 'OC', 'Envío', 'Estado', 'Observaciones']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border

    for row, oc in enumerate(ocs, 2):
        valores = [
            oc.id,
            oc.fecha.strftime('%d/%m/%Y') if oc.fecha else '',
            oc.empresa.nombre if oc.empresa else '',
            oc.oc or '',
            oc.envio,
            oc.estado,
            oc.observaciones or ''
        ]
        fill = PatternFill('solid', fgColor='EBF3FB') if row % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=10)
            c.border = border
            c.fill = fill

    anchos = [6, 12, 30, 15, 18, 25, 40]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="OC_Ventas.xlsx"'}
    )