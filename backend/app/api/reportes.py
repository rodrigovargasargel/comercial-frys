from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.db.database import get_db
from app.models.produccion import (
    OrdenProduccion, ProduccionExtrusora, DetalleProduccionExtrusora
)
from app.models.selladora import (
    OPSelladora, ProduccionSelladora, ProduccionSelladoraDetalle
)

router = APIRouter(prefix="/reportes", tags=["Reportes"])


def get_densidades_selladora(db, op_ids):
    from app.models.selladora import ProduccionSelladora, ProduccionSelladoraDetalle
    from app.models.produccion import OrdenProduccion, ProduccionExtrusora, DetalleProduccionExtrusora
    
    if not op_ids:
        return {}
    
    rows = db.query(
        ProduccionSelladora.op_id,
        OrdenProduccion.densidad
    ).join(
        ProduccionSelladoraDetalle,
        ProduccionSelladoraDetalle.produccion_selladora_id == ProduccionSelladora.id
    ).join(
        DetalleProduccionExtrusora,
        DetalleProduccionExtrusora.id == ProduccionSelladoraDetalle.detalle_extrusora_id
    ).join(
        ProduccionExtrusora,
        ProduccionExtrusora.id == DetalleProduccionExtrusora.produccion_extrusora_id
    ).join(
        OrdenProduccion,
        OrdenProduccion.id == ProduccionExtrusora.op_id
    ).filter(
        ProduccionSelladora.op_id.in_(op_ids)
    ).distinct().all()
    
    return {r.op_id: 'AD' if r.densidad == 'alta' else 'BD' for r in rows}


def get_lunes(fecha: date) -> date:
    return fecha - timedelta(days=fecha.weekday())


def get_semana_datos(db: Session, lunes: date):
    viernes = lunes + timedelta(days=4)
    dias = [lunes + timedelta(days=i) for i in range(5)]


      



    # --- EXTRUSORA: kg por producto/color/ancho/espesor/densidad por fecha y turno ---
    ext_rows = db.query(
        OrdenProduccion.producto_id,
        OrdenProduccion.color_id,
        OrdenProduccion.ancho,
        OrdenProduccion.espesor,
        OrdenProduccion.densidad,
        ProduccionExtrusora.fecha,
        ProduccionExtrusora.turno,
        func.sum(DetalleProduccionExtrusora.kg).label('total_kg')
    ).join(ProduccionExtrusora, ProduccionExtrusora.op_id == OrdenProduccion.id)\
     .join(DetalleProduccionExtrusora, DetalleProduccionExtrusora.produccion_extrusora_id == ProduccionExtrusora.id)\
     .filter(ProduccionExtrusora.fecha >= lunes, ProduccionExtrusora.fecha <= viernes)\
     .group_by(
         OrdenProduccion.producto_id, OrdenProduccion.color_id,
         OrdenProduccion.ancho, OrdenProduccion.espesor, OrdenProduccion.densidad,
         ProduccionExtrusora.fecha, ProduccionExtrusora.turno
     ).all()

    # --- SELLADORA: unidades por producto/color/ancho/espesor por fecha y turno ---
    sell_rows = db.query(
    OPSelladora.id,
    OPSelladora.producto_id,
    OPSelladora.color_id,
    OPSelladora.ancho,
    OPSelladora.espesor,
    OPSelladora.largo,  # ← agregar aquí
    ProduccionSelladora.fecha,
    ProduccionSelladora.turno,
    func.sum(ProduccionSelladoraDetalle.unidades).label('total_unidades')
        ).join(ProduccionSelladora, ProduccionSelladora.op_id == OPSelladora.id)\
        .join(ProduccionSelladoraDetalle, ProduccionSelladoraDetalle.produccion_selladora_id == ProduccionSelladora.id)\
        .filter(ProduccionSelladora.fecha >= lunes, ProduccionSelladora.fecha <= viernes)\
        .group_by(
            OPSelladora.id, OPSelladora.producto_id, OPSelladora.color_id,
            OPSelladora.ancho, OPSelladora.espesor, OPSelladora.largo,
            ProduccionSelladora.fecha, ProduccionSelladora.turno
        ).all()

    

    return dias, ext_rows, sell_rows


@router.get("/semana")
def get_reporte_semana(fecha: Optional[str] = None, db: Session = Depends(get_db)):
    from app.models.producto import Producto
    from app.models.produccion import Color

    hoy = date.today()
    base = date.fromisoformat(fecha) if fecha else hoy
    lunes = get_lunes(base)
    dias, ext_rows, sell_rows = get_semana_datos(db, lunes)
    viernes = lunes + timedelta(days=4)  # ← agregar esta línea

    # Construir estructura de filas
    productos_cache = {}
    colores_cache = {}

    def get_producto(pid):
        if pid not in productos_cache:
            p = db.query(Producto).filter(Producto.id == pid).first()
            productos_cache[pid] = p.nombre if p else str(pid)
        return productos_cache[pid]

    def get_color(cid):
        if cid not in colores_cache:
            c = db.query(Color).filter(Color.id == cid).first()
            colores_cache[cid] = c.nombre if c else str(cid)
        return colores_cache[cid]

    # Agrupar extrusora
    ext_data = {}
    for r in ext_rows:
        dens = 'AD' if r.densidad == 'alta' else 'BD'
        key = f"EXT|{get_producto(r.producto_id)}|{dens}|{get_color(r.color_id)}|{r.ancho}x{r.espesor}"
        if key not in ext_data:
            ext_data[key] = {'label': f"KG {get_producto(r.producto_id)} {dens} {get_color(r.color_id)} {r.ancho}x{r.espesor} ", 'dia': {}, 'noche': {}}
        fecha_str = r.fecha.isoformat()
        ext_data[key][r.turno][fecha_str] = round(float(r.total_kg), 2)


   # Obtener densidades de selladora
    op_sell_ids = list(set(r.producto_id for r in sell_rows))  # ids únicos
    # Mejor usar op_ids reales
    op_sell_ids_real = db.query(OPSelladora.id)\
        .join(ProduccionSelladora, ProduccionSelladora.op_id == OPSelladora.id)\
        .filter(ProduccionSelladora.fecha >= lunes, ProduccionSelladora.fecha <= viernes)\
        .distinct().all()
    op_sell_ids_list = [r.id for r in op_sell_ids_real]
    densidades_sell = get_densidades_selladora(db, op_sell_ids_list)       

   # Agrupar selladora
    sell_data = {}
    for r in sell_rows:
        # Buscar densidad de esta OP selladora
        op_sell = db.query(OPSelladora).filter(
            OPSelladora.producto_id == r.producto_id,
            OPSelladora.color_id == r.color_id,
            OPSelladora.ancho == r.ancho,
            OPSelladora.espesor == r.espesor,
            OPSelladora.largo == r.largo
        ).first()
        #dens = densidades_sell.get(op_sell.id, '') if op_sell else ''
        dens = densidades_sell.get(r.id, '')
        key = f"SELL|{get_producto(r.producto_id)}|{get_color(r.color_id)}|{r.ancho}x{r.espesor}x{r.largo}|{dens}"
        if key not in sell_data:
            sell_data[key] = {
                'label': f"UN {get_producto(r.producto_id)} {dens} {get_color(r.color_id)} {int(r.ancho)}x{int(r.largo)}x{int(r.espesor)}",
                'dia': {}, 'noche': {}
            }
        fecha_str = r.fecha.isoformat()
        sell_data[key][r.turno][fecha_str] = int(r.total_unidades)

    return {
        'lunes': lunes.isoformat(),
        'viernes': (lunes + timedelta(days=4)).isoformat(),
        'dias': [d.isoformat() for d in dias],
        'extrusora': list(ext_data.values()),
        'selladora': list(sell_data.values())
    }


@router.get("/semana/excel")
def descargar_excel_semana(fecha: Optional[str] = None, db: Session = Depends(get_db)):
    hoy = date.today()
    base = date.fromisoformat(fecha) if fecha else hoy
    lunes = get_lunes(base)
    dias, ext_rows, sell_rows = get_semana_datos(db, lunes)
    viernes = lunes + timedelta(days=4)  # ← agregar

    from app.models.producto import Producto
    from app.models.produccion import Color

    productos_cache = {}
    colores_cache = {}

    def get_producto(pid):
        if pid not in productos_cache:
            p = db.query(Producto).filter(Producto.id == pid).first()
            productos_cache[pid] = p.nombre if p else str(pid)
        return productos_cache[pid]

    def get_color(cid):
        if cid not in colores_cache:
            c = db.query(Color).filter(Color.id == cid).first()
            colores_cache[cid] = c.nombre if c else str(cid)
        return colores_cache[cid]
    

    # Obtener densidades de selladora
    op_sell_ids_real = db.query(OPSelladora.id)\
        .join(ProduccionSelladora, ProduccionSelladora.op_id == OPSelladora.id)\
        .filter(ProduccionSelladora.fecha >= lunes, ProduccionSelladora.fecha <= viernes)\
        .distinct().all()
    op_sell_ids_list = [r.id for r in op_sell_ids_real]
    densidades_sell = get_densidades_selladora(db, op_sell_ids_list)


    wb = Workbook()
    ws = wb.active
    ws.title = f"Semana {lunes.strftime('%d-%m-%Y')}"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1F3864')
    dia_fill = PatternFill('solid', fgColor='2E75B6')
    noche_fill = PatternFill('solid', fgColor='1F3864')


def get_densidad_op_selladora(db, op_selladora_id):
    from app.models.selladora import ProduccionSelladora, ProduccionSelladoraDetalle
    from app.models.produccion import OrdenProduccion, ProduccionExtrusora, DetalleProduccionExtrusora
    
    det = db.query(ProduccionSelladoraDetalle)\
        .join(ProduccionSelladora, ProduccionSelladoraDetalle.produccion_selladora_id == ProduccionSelladora.id)\
        .filter(ProduccionSelladora.op_id == op_selladora_id)\
        .first()
    
    if det and det.detalle_extrusora and det.detalle_extrusora.produccion:
        op_ext = db.query(OrdenProduccion).filter(
            OrdenProduccion.id == det.detalle_extrusora.produccion.op_id
        ).first()
        if op_ext:
            return 'AD' if op_ext.densidad == 'alta' else 'BD'
    return ''  


# ************ REPORTE STOCK **************     

@router.get("/stock")
def get_stock(db: Session = Depends(get_db)):
    from app.models.producto import Producto
    from app.models.produccion import Color, OrdenProduccion, ProduccionExtrusora, DetalleProduccionExtrusora
    from app.models.selladora import OPSelladora, ProduccionSelladora, ProduccionSelladoraDetalle
    from sqlalchemy import func

    # --- EXTRUSORA: kg producidos por producto/color/ancho/espesor/densidad ---
    # --- EXTRUSORA ---
    ext_rows = db.query(
        OrdenProduccion.producto_id,
        OrdenProduccion.color_id,
        OrdenProduccion.ancho,
        OrdenProduccion.espesor,
        OrdenProduccion.densidad,
        func.sum(DetalleProduccionExtrusora.kg).label('kg_total'),
        func.count(DetalleProduccionExtrusora.id).label('nro_rollos')
    ).join(ProduccionExtrusora, ProduccionExtrusora.op_id == OrdenProduccion.id)\
    .join(DetalleProduccionExtrusora, DetalleProduccionExtrusora.produccion_extrusora_id == ProduccionExtrusora.id)\
    .group_by(
        OrdenProduccion.producto_id, OrdenProduccion.color_id,
        OrdenProduccion.ancho, OrdenProduccion.espesor, OrdenProduccion.densidad
    ).all()

    # --- SELLADORA: unidades producidas por producto/color/ancho/espesor/largo ---
    sell_rows = db.query(
        OPSelladora.id,
        OPSelladora.producto_id,
        OPSelladora.color_id,
        OPSelladora.ancho,
        OPSelladora.espesor,
        OPSelladora.largo,
        func.sum(ProduccionSelladoraDetalle.unidades).label('unidades_total')
    ).join(ProduccionSelladora, ProduccionSelladora.op_id == OPSelladora.id)\
     .join(ProduccionSelladoraDetalle, ProduccionSelladoraDetalle.produccion_selladora_id == ProduccionSelladora.id)\
     .group_by(
         OPSelladora.id, OPSelladora.producto_id, OPSelladora.color_id,
         OPSelladora.ancho, OPSelladora.espesor, OPSelladora.largo
     ).all()

     #************* contador saldo y rollos disponibles en stock

    from app.models.selladora import ProduccionSelladoraDetalle as PSD

        # kg usados en selladora por rollo de extrusora
    kg_usados_subq = db.query(
            DetalleProduccionExtrusora.id.label('det_id'),
            func.coalesce(func.sum(PSD.kilos), 0).label('kg_usado')
        ).outerjoin(PSD, PSD.detalle_extrusora_id == DetalleProduccionExtrusora.id)\
        .filter(PSD.es_pack_parcial == False)\
        .group_by(DetalleProduccionExtrusora.id).subquery()

        # rollos con saldo > 0 por grupo
    rollos_disponibles = db.query(
            OrdenProduccion.producto_id,
            OrdenProduccion.color_id,
            OrdenProduccion.ancho,
            OrdenProduccion.espesor,
            OrdenProduccion.densidad,
            func.count(DetalleProduccionExtrusora.id).label('rollos_disp'),
            func.sum(DetalleProduccionExtrusora.kg - func.coalesce(kg_usados_subq.c.kg_usado, 0)).label('kg_saldo')
        ).join(ProduccionExtrusora, ProduccionExtrusora.op_id == OrdenProduccion.id)\
        .join(DetalleProduccionExtrusora, DetalleProduccionExtrusora.produccion_extrusora_id == ProduccionExtrusora.id)\
        .outerjoin(kg_usados_subq, kg_usados_subq.c.det_id == DetalleProduccionExtrusora.id)\
        .filter((DetalleProduccionExtrusora.kg - func.coalesce(kg_usados_subq.c.kg_usado, 0)) > 0)\
        .group_by(
            OrdenProduccion.producto_id, OrdenProduccion.color_id,
            OrdenProduccion.ancho, OrdenProduccion.espesor, OrdenProduccion.densidad
        ).all()

        # Indexar por clave
    saldo_map = {
            (r.producto_id, r.color_id, r.ancho, r.espesor, r.densidad): {
                'rollos_disp': r.rollos_disp,
                'kg_saldo': round(float(r.kg_saldo), 2)
            } for r in rollos_disponibles
        }

     #************* hasta aca contador

    productos_cache = {}
    colores_cache = {}

    def get_producto(pid):
        if pid not in productos_cache:
            p = db.query(Producto).filter(Producto.id == pid).first()
            productos_cache[pid] = p.nombre if p else str(pid)
        return productos_cache[pid]

    def get_color(cid):
        if cid not in colores_cache:
            c = db.query(Color).filter(Color.id == cid).first()
            colores_cache[cid] = c.nombre if c else str(cid)
        return colores_cache[cid]

    # Densidades selladora
    op_sell_ids = [r.id for r in sell_rows]
    densidades_sell = get_densidades_selladora(db, op_sell_ids)

    extrusora = []
    for r in ext_rows:  
        dens = 'AD' if r.densidad == 'alta' else 'BD'
        key = (r.producto_id, r.color_id, r.ancho, r.espesor, r.densidad)
        saldo_info = saldo_map.get(key, {'rollos_disp': 0, 'kg_saldo': 0})
        extrusora.append({
            'producto_id': r.producto_id,
            'color_id': r.color_id,
            'label': f"{get_producto(r.producto_id)} {dens} {get_color(r.color_id)} {r.ancho}x{r.espesor}",
            'kg_total': round(float(r.kg_total), 2),
            'nro_rollos': r.nro_rollos,
            'rollos_disponibles': saldo_info['rollos_disp'],
            'kg_saldo': saldo_info['kg_saldo'],
            'densidad': r.densidad,
            'ancho': r.ancho,
            'espesor': r.espesor
        })

    selladora = []
    for r in sell_rows:
        dens = densidades_sell.get(r.id, '')
        selladora.append({
            'op_id': r.id,
            'producto_id': r.producto_id,
            'color_id': r.color_id,
            'label': f"{get_producto(r.producto_id)} {dens} {get_color(r.color_id)} {int(r.ancho)}x{int(r.espesor)}x{int(r.largo)}",
            'unidades_total': int(r.unidades_total),
            'ancho': r.ancho,
            'espesor': r.espesor,
            'largo': r.largo
        })

    return {
        'extrusora': extrusora,
        'selladora': selladora
    }

@router.get("/stock/trazabilidad-extrusora")
def trazabilidad_extrusora(producto_id: int, color_id: int, ancho: float, espesor: float, densidad: str, db: Session = Depends(get_db)):
    from app.models.produccion import OrdenProduccion, ProduccionExtrusora, DetalleProduccionExtrusora
    from app.models.selladora import ProduccionSelladora, ProduccionSelladoraDetalle, OPSelladora
    from app.models.producto import Producto
    from app.models.empresa import Empresa

    ops = db.query(OrdenProduccion).filter(
        OrdenProduccion.producto_id == producto_id,
        OrdenProduccion.color_id == color_id,
        OrdenProduccion.ancho == ancho,
        OrdenProduccion.espesor == espesor,
        OrdenProduccion.densidad == densidad
    ).all()

    result = []
    saldo = 0

    for op in ops:
        prods = db.query(ProduccionExtrusora).filter(ProduccionExtrusora.op_id == op.id).all()
        for prod in prods:
            dets = db.query(DetalleProduccionExtrusora)\
                .filter(DetalleProduccionExtrusora.produccion_extrusora_id == prod.id)\
                .order_by(DetalleProduccionExtrusora.numero_rollo).all()
            for det in dets:
                kg_usado_rollo = db.query(func.coalesce(func.sum(ProduccionSelladoraDetalle.kilos), 0))\
                    .filter(ProduccionSelladoraDetalle.detalle_extrusora_id == det.id)\
                    .filter(ProduccionSelladoraDetalle.es_pack_parcial == False).scalar() or 0
                kg_disponible_rollo = round(det.kg - float(kg_usado_rollo), 2)

                # ENTRADA
                saldo = round(saldo + det.kg, 2)
                result.append({
                    'fecha': prod.fecha.isoformat() if prod.fecha else '',
                    'es': 'E',
                    'lote': prod.lote,
                    'op_id': op.id,
                    'op_sell_id': None,
                    'cliente': '',
                    'numero_rollo': det.numero_rollo,
                    'kg': det.kg,
                    'kg_disponible_rollo': kg_disponible_rollo,
                    'producto_sellado': '',
                    'unidades_producidas': None,
                    'kg_ocupados': None,
                    'saldo_kg': saldo
                })

                # SALIDAS inmediatamente después de la entrada
                usos = db.query(ProduccionSelladoraDetalle)\
                    .filter(ProduccionSelladoraDetalle.detalle_extrusora_id == det.id)\
                    .filter(ProduccionSelladoraDetalle.es_pack_parcial == False).all()

                for uso in usos:
                    prod_sell = db.query(ProduccionSelladora).filter(
                        ProduccionSelladora.id == uso.produccion_selladora_id
                    ).first()
                    op_sell = db.query(OPSelladora).filter(
                        OPSelladora.id == prod_sell.op_id
                    ).first() if prod_sell else None
                    prod_nombre = ''
                    cliente_nombre = ''
                    op_sell_id = None
                    if op_sell:
                        op_sell_id = op_sell.id
                        if op_sell.producto_id:
                            p = db.query(Producto).filter(Producto.id == op_sell.producto_id).first()
                            prod_nombre = p.nombre if p else ''
                        if op_sell.empresa_id:
                            emp = db.query(Empresa).filter(Empresa.id == op_sell.empresa_id).first()
                            cliente_nombre = emp.nombre if emp else ''

                    saldo = round(saldo - uso.kilos, 2)
                    result.append({
                        'fecha': prod_sell.fecha.isoformat() if prod_sell and prod_sell.fecha else '',
                        'es': 'S',
                        'lote': prod.lote,
                        'op_id': op.id,
                        'op_sell_id': op_sell_id,
                        'cliente': cliente_nombre,
                        'numero_rollo': det.numero_rollo,
                        'kg': det.kg,
                        'kg_disponible_rollo': None,
                        'producto_sellado': prod_nombre,
                        'unidades_producidas': uso.unidades,
                        'kg_ocupados': uso.kilos,
                        'saldo_kg': saldo
                    })

    return result

@router.get("/stock/trazabilidad-selladora")
def trazabilidad_selladora(op_id: int, db: Session = Depends(get_db)):
    from app.models.selladora import OPSelladora, ProduccionSelladora, ProduccionSelladoraDetalle
    from app.models.empresa import Empresa

    op = db.query(OPSelladora).filter(OPSelladora.id == op_id).first()
    if not op:
        return []

    prods = db.query(ProduccionSelladora).filter(ProduccionSelladora.op_id == op_id).all()

    result = []
    saldo_acumulado = 0

    for prod in prods:
        dets = db.query(ProduccionSelladoraDetalle)\
            .filter(ProduccionSelladoraDetalle.produccion_selladora_id == prod.id).all()
        for det in dets:
            saldo_acumulado += det.unidades
            cliente = ''
            if op.empresa_id:
                emp = db.query(Empresa).filter(Empresa.id == op.empresa_id).first()
                cliente = emp.nombre if emp else ''

            lote = ''
            if det.detalle_extrusora and det.detalle_extrusora.produccion:
                lote = det.detalle_extrusora.produccion.lote

            result.append({
                'fecha': prod.fecha.isoformat() if prod.fecha else '',
                'lote': lote,
                'es': 'E',
                'cantidad': det.unidades,
                'cliente': cliente,
                'saldo': saldo_acumulado
            })

    return result

    