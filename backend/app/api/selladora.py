from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List
from app.db.database import get_db
from app.schemas.selladora import (
    OPSelladoraCreate, OPSelladoraUpdate, OPSelladoraOut,
    ProduccionSelladoraCreate, ProduccionSelladoraOut,
    ProduccionSelladoraDetalleCreate, ProduccionSelladoraDetalleOut,
    DetalleExtrusoraDisponible
)
from app.services import selladora_service
from app.models.maquina import Maquina, TipoMaquina
from app.schemas.materia_prima import MPTipoOut

from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as date_type

router = APIRouter(prefix="/selladora", tags=["Selladora"])

@router.get("/maquinas-selladoras")
def listar_maquinas_selladoras(db: Session = Depends(get_db)):
    tipo = db.query(TipoMaquina).filter(TipoMaquina.nombre == "Selladora").first()
    if not tipo:
        return []
    return db.query(Maquina).filter(Maquina.tipo_maquina_id == tipo.id).all()

@router.get("/ops", response_model=List[OPSelladoraOut])
def listar_ops(db: Session = Depends(get_db)):
    return selladora_service.get_all_ops(db)

@router.post("/ops", response_model=OPSelladoraOut, status_code=201)
def crear_op(data: OPSelladoraCreate, db: Session = Depends(get_db)):
    return selladora_service.create_op(db, data)

@router.put("/ops/{op_id}", response_model=OPSelladoraOut)
def actualizar_op(op_id: int, data: OPSelladoraUpdate, db: Session = Depends(get_db)):
    op = selladora_service.update_op(db, op_id, data)
    if not op:
        raise HTTPException(status_code=404, detail="OP no encontrada")
    return op

@router.delete("/ops/{op_id}", status_code=204)
def eliminar_op(op_id: int, db: Session = Depends(get_db)):
    if not selladora_service.delete_op(db, op_id):
        raise HTTPException(status_code=404, detail="OP no encontrada")

@router.get("/ops/{op_id}/producciones", response_model=List[ProduccionSelladoraOut])
def listar_producciones(op_id: int, db: Session = Depends(get_db)):
    return selladora_service.get_producciones_by_op(db, op_id)

@router.post("/producciones", response_model=ProduccionSelladoraOut, status_code=201)
def crear_produccion(data: ProduccionSelladoraCreate, db: Session = Depends(get_db)):
    return selladora_service.create_produccion(db, data)

@router.delete("/producciones/{prod_id}", status_code=204)
def eliminar_produccion(prod_id: int, db: Session = Depends(get_db)):
    if not selladora_service.delete_produccion(db, prod_id):
        raise HTTPException(status_code=404, detail="No encontrado")

@router.get("/producciones/{prod_id}/detalles", response_model=List[ProduccionSelladoraDetalleOut])
def listar_detalles(prod_id: int, db: Session = Depends(get_db)):
    from app.models.produccion import OrdenProduccion
    detalles = selladora_service.get_detalles_by_produccion(db, prod_id)
    result = []
    for d in detalles:
        lote = None
        fecha = None
        densidad = None
        if d.detalle_extrusora and d.detalle_extrusora.produccion:
            lote = d.detalle_extrusora.produccion.lote
            fecha = d.detalle_extrusora.produccion.fecha
            op_ext = db.query(OrdenProduccion).filter(
                OrdenProduccion.id == d.detalle_extrusora.produccion.op_id
            ).first()
            if op_ext:
                densidad = op_ext.densidad
            result.append(ProduccionSelladoraDetalleOut(
                id=d.id,
                detalle_extrusora_id=d.detalle_extrusora_id,
                q_paquetes=d.q_paquetes,
                q_unidades_por_paquete=d.q_unidades_por_paquete,
                unidades=d.unidades,
                kilos=d.kilos,
                kg_rollo_original=d.detalle_extrusora.kg if d.detalle_extrusora else None,  # ← este
                imprimir_kg=d.imprimir_kg,
                numero_rollo=d.detalle_extrusora.numero_rollo if d.detalle_extrusora else 0,
                lote_extrusora=lote,
                fecha_extrusora=fecha,
                densidad_extrusora=densidad,
                created_at=d.created_at
            ))
    return result

@router.get("/rollos-disponibles", response_model=List[DetalleExtrusoraDisponible])
def rollos_disponibles(color_id: int, ancho: float, espesor: float, db: Session = Depends(get_db)):
    rollos = selladora_service.get_rollos_disponibles(db, color_id, ancho, espesor)
    return [
        DetalleExtrusoraDisponible(
            id=r.id,
            numero_rollo=r.numero_rollo,
            kg=r.kg,
            lote=r.produccion.lote if r.produccion else '',
            fecha_produccion=r.produccion.fecha if r.produccion else None
        ) for r in rollos
    ]

@router.post("/detalles", response_model=ProduccionSelladoraDetalleOut, status_code=201)
def crear_detalle(data: ProduccionSelladoraDetalleCreate, db: Session = Depends(get_db)):
    try:
        return selladora_service.create_detalle(db, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/detalles/{detalle_id}", status_code=204)
def eliminar_detalle(detalle_id: int, db: Session = Depends(get_db)):
    if not selladora_service.delete_detalle(db, detalle_id):
        raise HTTPException(status_code=404, detail="No encontrado")

@router.get("/productos-selladora")
def listar_productos_selladora(db: Session = Depends(get_db)):
    from app.models.producto import Producto
    from app.models.maquina import TipoMaquina
    tipo = db.query(TipoMaquina).filter(TipoMaquina.nombre == "Selladora").first()
    if not tipo:
        return []
    return db.query(Producto).filter(
    Producto.tipo_maquina_id == tipo.id
).all()     

@router.post("/ops/{op_id}/trazabilidad")
def generar_trazabilidad(op_id: int, body: dict, db: Session = Depends(get_db)):
    from app.models.produccion import OrdenProduccion, ProduccionExtrusora, DetalleProduccionExtrusora
    from app.models.selladora import ProduccionSelladoraDetalle

    op = selladora_service.get_op_by_id(db, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="OP no encontrada")

    cliente = body.get('cliente', '')
    nro_guia = body.get('nro_guia', '')
    oc_cliente = body.get('oc_cliente', '')
    # Obtener todos los detalles de selladora de esta OP
    prods = selladora_service.get_producciones_by_op(db, op_id)
    
    
    thin = Side(style='thin')
    medium = Side(style='medium')
    border_full = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1F3864')
    subheader_fill = PatternFill('solid', fgColor='2E75B6')
    row_fill_alt = PatternFill('solid', fgColor='EBF3FB')

    wb = Workbook()
    ws = wb.active
    ws.title = str(op_id)

    # Anchos columnas A-K
    anchos = [8, 14, 36, 14, 14, 12, 14, 12, 10, 10, 12]
    for i, w in enumerate(anchos, 1):
     ws.column_dimensions[get_column_letter(i)].width = w

    hoy = date_type.today()
    fecha_str = hoy.strftime('%d/%m/%Y')

    fecha_fact_raw = body.get('fecha_fact', '')
    if fecha_fact_raw and '-' in fecha_fact_raw:
        partes = fecha_fact_raw.split('-')
        fecha_fact_str = f"{partes[2]}/{partes[1]}/{partes[0]}"
    else:
        fecha_fact_str = fecha_str

    # Nº
    ws['H4'] = 'Nº :'
    ws['H4'].font = Font(bold=True, size=10, color='1F3864')
    ws['I4'] = op_id
    ws['I4'].font = Font(bold=True, size=12, color='1F3864')
    ws['I4'].alignment = Alignment(horizontal='center')

    ws['H3'] = 'N° de Informe :'
    ws['H3'].font = Font(bold=True, size=10, color='1F3864')
    numero_informe = f"{hoy.strftime('%d%m%y')}"
    ws['I3'] = numero_informe
    ws['I3'].font = Font(bold=True, size=10, color='1F3864')
    ws['I3'].alignment = Alignment(horizontal='center')

    # Empresa
    empresa_rows = [
        (4, 'COMERCIALIZADORA Y DISTRIBUIDORA FRYS LTDA', True, 12),
        (5, 'RUT : 76386703-K', False, 10),
        (6, 'CAM TEPUAL KM 3 P APIAS MONTT P-9', False, 10),
        (7, 'PUERTO MONTT - CHILE', False, 10),
        (8, 'TEL (065) 2254554 - (09) 96797817', False, 10),
        (9, 'GIRO : COMERCIALIZACION DE INSUMOS Y PRODUCTOS INDUSTRIALES', False, 10),
        (10, 'rmolina@comercialfrys.cl', False, 10),
    ]
    for row, val, bold, size in empresa_rows:
        c = ws[f'A{row}']
        c.value = val
        c.font = Font(bold=bold, size=size, color='1F3864')

    # Título TRAZABILIDAD
    ws.merge_cells('A12:J12')
    c = ws['A12']
    c.value = 'TRAZABILIDAD'
    c.font = Font(bold=True, size=22, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = header_fill
    ws.row_dimensions[12].height = 22

    # A/TO y FECHA
    ws['A13'] = 'A/TO :'
    ws['A13'].font = Font(bold=True, size=10, color='1F3864')
    ws['C13'] = cliente.upper()
    ws['C13'].font = Font(bold=True, size=11)
    ws['H13'] = 'FECHA/DATE :'
    ws['H13'].font = Font(bold=True, size=10, color='1F3864')
    ws['J13'] = fecha_str
    ws['J13'].font = Font(size=10)
    ws['J13'].alignment = Alignment(horizontal='center')

    # Recopilar lotes únicos de los rollos usados
    lotes = set()
    for prod in prods:
        dets = db.query(ProduccionSelladoraDetalle)\
            .filter(ProduccionSelladoraDetalle.produccion_selladora_id == prod.id).all()
        for d in dets:
            if d.detalle_extrusora and d.detalle_extrusora.produccion:
                lotes.add(str(d.detalle_extrusora.produccion.lote))

    ws['A14'] = 'ORDEN PEDIDO MP :'
    ws['A14'].font = Font(bold=True, size=10, color='1F3864')
    ws['C14'] = ' '.join(sorted(lotes))
    ws['C14'].font = Font(size=10)

    ws['A15'] = 'ORDEN DE COMPRA'
    ws['A15'].font = Font(bold=True, size=10, color='1F3864')
    ws['C15'] = oc_cliente if oc_cliente else ''
    ws['C15'].font = Font(bold=True, size=10)

    # Encabezado tabla — sin COD ni PROVEEDOR
    HEADER_ROW = 16
    headers = [
        ('A','GUIA/FACT'), ('B','FECHA'), ('C','NOMBRE PRODUCTO'),
        ('E','F. FABRICACIÓN'), ('F','LOTE'), ('G','F. VENCIMIENTO'),
        ('H','N° ROLLO'), ('I','N° PQTE'), ('J','UNID'), ('K','T. UNID')
    ]
    ws.row_dimensions[HEADER_ROW].height = 16
    for col, h in headers:  
        c = ws[f'{col}{HEADER_ROW}']
        c.value = h
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = subheader_fill
        c.border = border_full
        c.alignment = Alignment(horizontal='center', vertical='center')

   # Merge columnas C y D para Nombre Producto
    ws.merge_cells(f'C{HEADER_ROW}:D{HEADER_ROW}')
    densidad_str = ''
    for prod in prods:
        dets = db.query(ProduccionSelladoraDetalle)\
            .filter(ProduccionSelladoraDetalle.produccion_selladora_id == prod.id).first()
        if dets and dets.detalle_extrusora and dets.detalle_extrusora.produccion:
            from app.models.produccion import OrdenProduccion
            op_ext = db.query(OrdenProduccion).filter(
                OrdenProduccion.id == dets.detalle_extrusora.produccion.op_id
            ).first()
            if op_ext:
                densidad_str = 'AD' if op_ext.densidad == 'alta' else 'BD'
                break

    
    # Nombre producto
    nombre_producto = f"{op.producto.nombre if op.producto else ''} {densidad_str} {op.color.nombre if op.color else ''} {int(op.ancho)}x{int(op.largo)}x{int(op.espesor)}"

    DATA_START = HEADER_ROW + 1
    fila = DATA_START

    for prod in prods:
        dets = db.query(ProduccionSelladoraDetalle)\
            .filter(ProduccionSelladoraDetalle.produccion_selladora_id == prod.id)\
            .all()
        for d in dets:
            fill = PatternFill('solid', fgColor='FFFFFF') if fila % 2 == 0 else row_fill_alt
            ws.row_dimensions[fila].height = 15

            # Fechas desde el rollo de extrusora
            fecha_fab = ''
            fecha_venc = ''
            lote_rollo = ''
            if d.detalle_extrusora and d.detalle_extrusora.produccion:
                fecha_prod = d.detalle_extrusora.produccion.fecha
                if fecha_prod:
                    fecha_fab = fecha_prod.strftime('%d/%m/%Y')
                    from datetime import timedelta
                    fecha_venc_dt = date_type(fecha_prod.year + 1, fecha_prod.month, fecha_prod.day)
                    fecha_venc = fecha_venc_dt.strftime('%d/%m/%Y')
                lote_rollo = str(d.detalle_extrusora.produccion.lote)

            numero_rollo = d.detalle_extrusora.numero_rollo if d.detalle_extrusora else ''

            valores = [
                ('A', nro_guia),
                ('B', fecha_fact_str),
                ('C', nombre_producto),
                ('E', fecha_fab),
                ('F', lote_rollo),
                ('G', fecha_venc),
                ('H', numero_rollo),
                ('I', d.q_paquetes),
                ('J', d.q_unidades_por_paquete),
                ('K', f'=J{fila}*I{fila}'),
            ]
            for col, val in valores:
                c = ws[f'{col}{fila}']
                c.value = val
                c.font = Font(size=10)
                c.border = border_full
                c.fill = fill
                if col in ('H', 'I', 'J', 'K'):
                    c.alignment = Alignment(horizontal='center')

            # Merge C y D para nombre producto en cada fila
            ws.merge_cells(f'C{fila}:D{fila}')
            fila += 1

    # Total
    total_row = fila
    ws.merge_cells(f'A{total_row}:J{total_row}')
    c = ws[f'A{total_row}']
    c.value = 'TOTAL'
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = header_fill
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = border_full
    tc = ws[f'K{total_row}']
    tc.value = f'=SUM(K{DATA_START}:K{fila-1})'
    tc.font = Font(bold=True, size=11, color='FFFFFF')
    tc.fill = header_fill
    tc.border = border_full
    tc.alignment = Alignment(horizontal='center', vertical='center')

    # Firma
    firma_row = total_row + 4
    for i, linea in enumerate(['Renzo Molina E.', 'Dpto. Comercial', 'F : 09-96797817', 'rmolina@comercialfrys.cl']):
        c = ws[f'D{firma_row+i}']
        c.value = linea
        c.font = Font(size=10, color='1F3864', italic=(i > 0))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Trazabilidad-OP{op_id}-{cliente}.xlsx"
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )